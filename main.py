# -*- coding: utf-8 -*-
"""
MIT License

Copyright (c) 2023 Mark Shuck

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

@author: Mark Shuck
email: mark@shuck.engineering
"""

import numpy as np
import pandas as pd
import datetime
import scipy.optimize as optimize

import UnitConversion as Unc
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

uc = Unc.UnitConvert()

ver = 'v0.3.0'

degree_sign = u'\N{DEGREE SIGN}'

#todo- change dataFrame structure to remove first row as the "config" setup and put back in to dedicated dataFrame
# remove other references to df_N, df_E, df_L and add in calculation type column "Normal/Emergency/Load Dump"
# add input verification that input dataframe contains enough & proper data to perform all calculations.


class IEEE738:
    true_to_standard = True
    conductor_temp_steps = 6

    # def __init__(self):


    @staticmethod
    def import_config(path, sheet_name):
        """
        Imports configuration from Excel file
        :param path: path to Excel file
        :param sheet_name: Excel sheet name with list of configurations (nonconductor based parameters for calculations)
        :return: Configurations and parameters listed in file (pandas dataframe)
        """
        config_list = pd.read_excel(io=path, sheet_name=sheet_name, engine='openpyxl')
        return config_list

    @staticmethod
    def import_conductor(path, sheet_name):
        """
        Imports list of conductors & corresponding parameters and conductor specifications (max temp) from Excel file
        and sorts data smallest to largest and A-Z to be used later on
        :param path: path to file
        :param sheet_name: Excel sheet name with list of conductor parameters
        :return: Conductors and parameters listed in file (pandas dataframe)
        """
        # read in list of conductors and corresponding parameters
        conductor_list = pd.read_excel(io=path, sheet_name=sheet_name[0], engine='openpyxl')
        # read temperature ranges for the different conductor specifications (ACCC/ASCR/etc...)
        conductor_spec = pd.read_excel(io=path, sheet_name=sheet_name[1], engine='openpyxl')
        # sort conductor dataframe based on conductor outer diameter
        conductor_list.sort_values('Metal OD', ascending=True, inplace=True)
        # sort conductor spec A-Z
        conductor_spec.sort_values('Conductor Spec', ascending=True, inplace=True)
        return conductor_list, conductor_spec

    @staticmethod
    def add_calc_columns(df):
        data = ['qc heat loss', 'qc0', 'qc1', 'qc2', 'uf', 'kf', 'pf', 'Qse', 'theta', 'hc: solar altitude', 'delta',
                'omega', 'chi', 'qs heat gain', 'solar altitude correction factor', 'qr heat loss', 'day of year',
                'k angle', 'solar azimuth constant', 'solar azimuth', 'conductor temperature']
        df = pd.concat([df.reset_index(drop=True), pd.DataFrame(columns=data)], axis=1)
        return df

    @staticmethod
    def unit_conversion(df_conductor, df_spec, df_config):

        unit_selection = None

        df_conductor_wind_list = pd.DataFrame()
        df_conductor_wind_angle_list = pd.DataFrame()
        df_conductor_length_list = pd.DataFrame()
        df_conductor_temp_list = pd.DataFrame()
        df_conductor_spec_temp_list = pd.DataFrame()
        df_config_temp_list = pd.DataFrame()
        df_config_length_list = pd.DataFrame()

        df_conductor_adjusted = pd.DataFrame(df_conductor)
        df_config_adjusted = pd.DataFrame(df_config)

        calculation_units = df_config.at[0, 'calculation units']

        if uc.units_lookup[calculation_units] == uc.metric_value:
            unit_selection = 2
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            unit_selection = 3

        conductor_wind_list = (
            ('normal wind speed', 'normal wind speed units', 'm/s', 'ft/hr'),
            ('emergency wind speed', 'emergency wind speed units', 'm/s', 'ft/hr'),
        )
        conductor_wind_angle_list = (
            ('normal wind angle', 'normal wind angle units', 'deg', 'deg'),
            ('emergency wind angle', 'emergency wind angle units', 'deg', 'deg')
        )
        conductor_length_list = (
            ('Cond Wire Diameter', 'Cond Wire Diameter Units', 'mm', 'in'),
            ('Core Wire Diameter', 'Core Wire Diameter Units', 'mm', 'in'),
            ('Core OD', 'Core OD Units', 'mm', 'in'),
            ('Metal OD', 'Metal OD Units', 'mm', 'in'),
            ('resistance distance', 'resistance distance units', 'm', 'ft')
        )
        conductor_temp_list = (
            ('low resistance temperature', 'resistance temperature units', 'C', 'C'),
            ('high resistance temperature', 'resistance temperature units', 'C', 'C')
        )
        conductor_spec_temp_list = (
            ('normal temperature rating', 'normal temperature rating units', 'C', 'C'),
            ('emergency temperature rating', 'emergency temperature rating units', 'C', 'C')
        )
        config_temp_list = (
            ('ambient air temperature lower range', 'ambient air temperature units', 'C', 'C'),
            ('ambient air temperature upper range', 'ambient air temperature units', 'C', 'C'),
            ('temperature increment', 'ambient air temperature units', 'C', 'C'),
            ('ambient air temperature', 'ambient air temperature units', 'C', 'C')
        )
        config_length_list = (
            ('elevation', 'elevation units', 'm', 'ft'),
        )

        for x in conductor_wind_list:
            value = uc.speed_convert(df_config.at[0, x[0]], df_config.at[0, x[1]], x[unit_selection])
            df_conductor_wind_list.at[0, x[0]] = value
            df_conductor_wind_list.at[0, x[1]] = x[unit_selection]
            df_config_adjusted.drop(columns=x[0], axis=1, inplace=True)
            df_config_adjusted.drop(columns=x[1], axis=1, inplace=True)

        for x in conductor_wind_angle_list:
            value = uc.angle_convert(df_config.at[0, x[0]], df_config.at[0, x[1]], x[unit_selection])
            df_conductor_wind_angle_list.at[0, x[0]] = value
            df_conductor_wind_angle_list.at[0, x[1]] = x[unit_selection]
            df_config_adjusted.drop(columns=x[0], axis=1, inplace=True)
            df_config_adjusted.drop(columns=x[1], axis=1, inplace=True)

        for x in conductor_length_list:
            value = uc.length_convert(df_conductor.at[0, x[0]], df_conductor.at[0, x[1]], x[unit_selection])
            df_conductor_length_list.at[0, x[0]] = value
            df_conductor_length_list.at[0, x[1]] = x[unit_selection]
            df_conductor_adjusted.drop(columns=x[0], axis=1, inplace=True)
            df_conductor_adjusted.drop(columns=x[1], axis=1, inplace=True)

        for x in conductor_temp_list:
            value = uc.temp_convert(df_conductor.at[0, x[0]], df_conductor.at[0, x[1]], 'C')
            df_conductor_temp_list.at[0, x[0]] = value
            df_conductor_temp_list.at[0, x[1]] = x[unit_selection]
            df_conductor_adjusted.drop(columns=x[0], axis=1, inplace=True)
            try:
                df_conductor_adjusted.drop(columns=x[1], axis=1, inplace=True)
            except KeyError:
                None

        for x in conductor_spec_temp_list:
            value = uc.temp_convert(df_spec.at[0, x[0]], df_spec.at[0, x[1]], 'C')
            df_conductor_spec_temp_list.at[0, x[0]] = value
            df_conductor_spec_temp_list.at[0, x[1]] = x[unit_selection]

        for x in config_temp_list:
            value = uc.temp_convert(df_config.at[0, x[0]], df_config.at[0, x[1]], 'C')
            df_config_temp_list.at[0, x[0]] = value
            df_config_temp_list.at[0, x[1]] = x[unit_selection]
            df_config_adjusted.drop(columns=x[0], axis=1, inplace=True)
            try:
                df_config_adjusted.drop(columns=x[1], axis=1, inplace=True)
            except KeyError:
                None

        for x in config_length_list:
            value = uc.length_convert(df_config.at[0, x[0]], df_config.at[0, x[1]], x[unit_selection])
            df_config_length_list.at[0, x[0]] = value
            df_config_length_list.at[0, x[1]] = x[unit_selection]
            df_config_adjusted.drop(columns=x[0], axis=1, inplace=True)
            df_config_adjusted.drop(columns=x[1], axis=1, inplace=True)

        df_conductor_adjusted = pd.concat([df_conductor_adjusted, df_conductor_wind_list.reset_index(drop=True),
                                           df_conductor_wind_angle_list.reset_index(drop=True),
                                           df_conductor_length_list.reset_index(drop=True),
                                           df_conductor_temp_list.reset_index(drop=True),
                                           df_conductor_spec_temp_list.reset_index(drop=True),
                                           df_config_temp_list.reset_index(drop=True),
                                           df_config_length_list.reset_index(drop=True)], axis=1)

        df_adjusted = pd.concat(
            [df_conductor_adjusted.reset_index(drop=True), df_config_adjusted.reset_index(drop=True)], axis=1)

        df_adjusted = df_adjusted.fillna(0)

        return df_adjusted

    def c_steady_state(self, df, calcType=None, _idx=None):
        # Configuration setup
        conductor_projection = None
        calculation_units = df.at[_idx, 'calculation units']
        elevation = df.at[_idx, 'elevation']
        emissivity = df.at[_idx, 'emissivity']
        solar_absorptivity = df.at[_idx, 'solar absorptivity']
        atmosphere = df.at[_idx, 'atmosphere']
        latitude = df.at[_idx, 'latitude']
        day = df.at[_idx, 'day']
        month = df.at[_idx, 'month']
        year = df.at[_idx, 'year']
        hour = df.at[_idx, 'hour']
        conductor_direction = df.at[_idx, 'conductor direction']
        ambient_air_temp = df.at[_idx, 'ambient air temperature']
        conductor_temperature = df.at[_idx, 'conductor temperature']

        # Conductor setup
        diameter = df.at[_idx, 'Metal OD']

        if uc.units_lookup[calculation_units] == uc.metric_value:
            conductor_projection = diameter / 1000
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            conductor_projection = diameter / 12

        d = {'high resistance Ω/unit': [df.at[0, 'high resistance Ω/unit']],  # high resistance
             'low resistance Ω/unit': [df['low resistance Ω/unit'].values[0]],  # low resistance
             'resistance temperature unit': [df['resistance temperature units'].values[0]],
             # resistance distance unit (C, F, K, R)
             'high resistance temperature': [df['high resistance temperature'].values[0]],
             # high resistance temp
             'low resistance temperature': [df['low resistance temperature'].values[0]],  # low resistance temp
             'resistance distance': [df['resistance distance'].values[0]],  # resistance distance
             'resistance distance unit': [df['resistance distance units'].values[0]]
             # resistance distance unit (mile/meter/etc...)
             }

        conductor_resistance = pd.DataFrame(d)
        # Normal Rating (wind speed and angle variable)

        if calcType == 'Emergency':
            conductor_wind_speed = df.at[_idx, 'emergency wind speed']
            conductor_wind_angle = df.at[_idx, 'emergency wind angle']
        else:
            conductor_wind_speed = df.at[_idx, 'normal wind speed']
            conductor_wind_angle = df.at[_idx, 'normal wind angle']

        current_rating_day, current_rating_night = self.c_SSRating(calculation_units, diameter, conductor_temperature,
                                                                   ambient_air_temp, elevation, conductor_wind_angle,
                                                                   conductor_wind_speed, emissivity, solar_absorptivity,
                                                                   atmosphere, latitude, day, month, year, hour,
                                                                   conductor_direction, conductor_projection,
                                                                   conductor_resistance, df, _idx)
        return current_rating_day, current_rating_night

    def c_load_dump(self, df, _idx):
        # Configuration setup
        conductor_projection = None
        calculation_units = df.at[_idx, 'calculation units']
        elevation = df.at[_idx, 'elevation']
        emissivity = df.at[_idx, 'emissivity']
        solar_absorptivity = df.at[_idx, 'solar absorptivity']
        atmosphere = df.at[_idx, 'atmosphere']
        latitude = df.at[_idx, 'latitude']
        day = df.at[_idx, 'day']
        month = df.at[_idx, 'month']
        year = df.at[_idx, 'year']
        hour = df.at[_idx, 'hour']
        conductor_direction = df.at[_idx, 'conductor direction']

        ambient_air_temp = df.at[_idx, 'ambient air temperature']

        conductor_temp_normal = df.at[_idx, 'normal temperature rating']
        conductor_temp_emergency = df.at[_idx, 'emergency temperature rating']
        conductor_wind_emergency = df.at[_idx, 'emergency wind speed']

        # Conductor setup
        diameter = df.at[_idx, 'Metal OD']

        if uc.units_lookup[calculation_units] == uc.metric_value:
            conductor_projection = diameter / 1000
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            conductor_projection = diameter / 12

        d = {'high resistance Ω/unit': [df['high resistance Ω/unit'].values[0]],  # high resistance
             'low resistance Ω/unit': [df['low resistance Ω/unit'].values[0]],  # low resistance
             'resistance temperature unit': [df['resistance temperature units'].values[0]],
             # resistance distance unit (C, F, K, R)
             'high resistance temperature': [df['high resistance temperature'].values[0]],
             # high resistance temp
             'low resistance temperature': [df['low resistance temperature'].values[0]],  # low resistance temp
             'resistance distance': [df['resistance distance'].values[0]],  # resistance distance
             'resistance distance unit': [df['resistance distance units'].values[0]]
             # resistance distance unit (mile/meter/etc...)
             }

        conductor_resistance = pd.DataFrame(d)

        wind_angle = df.at[_idx, 'emergency wind angle']

        mcp = self.c_mcp(df, _idx)

        duration = df.at[_idx, 'duration (minutes)']

        load_dump_day, load_dump_night = self.load_dump(calculation_units, diameter, conductor_temp_normal,
                                                        conductor_temp_emergency,
                                                        ambient_air_temp, elevation, wind_angle,
                                                        conductor_wind_emergency, emissivity, solar_absorptivity,
                                                        atmosphere, latitude, day, month, year, hour,
                                                        conductor_direction, conductor_projection,
                                                        conductor_resistance, mcp, duration)

        if df is not None:
            df.at[_idx, 'load dump rating daytime'] = load_dump_day
            df.at[_idx, 'load dump rating nighttime'] = load_dump_night
            df.at[_idx, 'load dump duration'] = duration

        return load_dump_day, load_dump_night

    def c_temperature_range(self, df_adjusted):
        """
        Generates a range of ambient temperatures (C) used for conductor rating calculations.
        Uses lower/upper/increment from configuration file to generate range.

        Generates a range of conductor temperatures (C) used for conductor rating calculations.
        Uses conductor_temp_steps & ambient increment to generate lower/upper bounds based on
        conductor normal/emergency ratings
        :param df_adjusted: data frame containing all configuration parameters adjusted to required units
        :return: two ranges
        1- ambient temperatures in degrees C
        2-conductor temperatures in degrees C
        """

        ambient_lower_range = df_adjusted.at[0, 'ambient air temperature lower range']
        ambient_upper_range = df_adjusted.at[0, 'ambient air temperature upper range']
        ambient_temp_inc = df_adjusted.at[0, 'temperature increment']
        conductor_normal_rating = df_adjusted.at[0, 'normal temperature rating']
        conductor_emergency_rating = df_adjusted.at[0, 'emergency temperature rating']

        if conductor_normal_rating == conductor_emergency_rating:
            temp_range_conductor = np.arange(conductor_normal_rating - self.conductor_temp_steps * ambient_temp_inc,
                                             conductor_normal_rating + (
                                                     1 + self.conductor_temp_steps) * ambient_temp_inc,
                                             ambient_temp_inc)
        else:
            temp_range_conductor = np.arange(conductor_normal_rating - self.conductor_temp_steps * ambient_temp_inc,
                                             conductor_emergency_rating + (
                                                     1 + self.conductor_temp_steps) * ambient_temp_inc,
                                             ambient_temp_inc)

        temp_range_ambient = np.arange(ambient_lower_range, ambient_upper_range + ambient_temp_inc, ambient_temp_inc)

        return temp_range_ambient, temp_range_conductor

    def c_reporting(self, df_conductor, df_spec, df_config):
        _idx = 1

        df_adjusted = self.unit_conversion(df_conductor, df_spec, df_config)

        df_adjusted = self.add_calc_columns(df_adjusted)

        temp_range_ambient, temp_range_conductor = self.c_temperature_range(df_adjusted)

        df = pd.DataFrame(df_adjusted)

        total_row = temp_range_ambient.size * temp_range_conductor.size + 1
        df_N = pd.concat([df] * total_row, axis=0, ignore_index=True)
        df_E = pd.concat([df] * total_row, axis=0, ignore_index=True)
        df_L = pd.concat([df] * total_row, axis=0, ignore_index=True)

        for i, element_i in enumerate(temp_range_ambient):
            for j, element_j in enumerate(temp_range_conductor):

                df_N.at[_idx, 'ambient air temperature'] = element_i
                df_N.at[_idx, 'conductor temperature'] = element_j
                df_E.at[_idx, 'ambient air temperature'] = element_i
                df_E.at[_idx, 'conductor temperature'] = element_j
                df_L.at[_idx, 'ambient air temperature'] = element_i
                df_L.at[_idx, 'conductor temperature'] = element_j

                _, _ = self.c_steady_state(df_N, 'Normal', _idx)
                _, _ = self.c_steady_state(df_E, 'Emergency', _idx)
                if j == 0:
                    _, _ = self.c_load_dump(df_L, _idx)
                else:
                    df_L.at[_idx, 'load dump rating daytime'] = df_L.at[_idx - 1, 'load dump rating daytime']
                    df_L.at[_idx, 'load dump rating nighttime'] = df_L.at[_idx - 1, 'load dump rating nighttime']
                _idx = _idx + 1

        # TODO add polynomial regression to replace nan with none zero.
        #  ex HD Copper 500 @ Tc = 55C & Amb = 40C I = nan
        #  mention this or highlight cell somehow

        # remove first row (configuration setup)
        # TODO --> likely a better way to make this work
        df_N = df_N.drop(labels=0, axis='index').reset_index(drop=False)
        df_E = df_E.drop(labels=0, axis='index').reset_index(drop=True)
        df_L = df_L.drop(labels=0, axis='index').reset_index(drop=True)

        return df_N, df_E, df_L

    @staticmethod
    def export_excel(df_n, df_e, df_l, df_config, filename_):
        wb = Workbook()
        ws = wb.active
        filename_ = filename_ + '.xlsx'
        ws.title = 'normal'
        for r in dataframe_to_rows(df_n, index=False, header=True):
            ws.append(r)

        ws = wb.create_sheet('emergency')
        for r in dataframe_to_rows(df_e, index=False, header=True):
            ws.append(r)

        ws = wb.create_sheet('load')
        for r in dataframe_to_rows(df_l, index=False, header=True):
            ws.append(r)

        ws = wb.create_sheet('config')
        for r in dataframe_to_rows(df_config, index=False, header=True):
            ws.append(r)

        wb.save(filename_)

    @staticmethod
    def current_steady_state(qr, qs, qc, r):
        """
        Calculates steady state current rating and returns results (Amps)
        :param qr: Radiated heat loss (W/m or W/ft)
        :param qs: Heat gain from The Sun (W/m or W/ft)
        :param qc: Convected heat loss (W/m or W/ft)
        :param r: Conductor resistance (ohms)
        :return: Steady state current rating (Amps)
        """
        if (qr - qs + qc) > 0:
            rating = np.sqrt((qr - qs + qc) / r)
        else:
            rating = 0
        return rating

    @staticmethod
    def c_uf(calculation_units, conductor_temp, ambient_air_temp, df=None, _idx=0):
        """
        Calculates dynamic viscosity of air and returns results (Pa-s or lb/ft-hr)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param conductor_temp: Conductor temperature (C)
        :param ambient_air_temp: Ambient air temperature (C)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Dynamic viscosity of air (Pa-s or lb/ft-hr)
        """
        uf = None
        t_film = (conductor_temp + ambient_air_temp) / 2

        if uc.units_lookup[calculation_units] == uc.metric_value:
            uf = (1.458 * 10 ** -6 * (t_film + 273.15) ** 1.5) / (t_film + 383.4)
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            uf = (0.00353 * (t_film + 273.15) ** 1.5) / (t_film + 383.4)
        if df is not None:
            df.at[_idx, 'uf'] = uf
        return uf

    @staticmethod
    def c_kf(calculation_units, conductor_temp, ambient_air_temp, df=None, _idx=0):
        """
        Calculates thermal conductivity of air and returns results (W/m*C or W/ft*C)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param conductor_temp: Conductor temperature (C)
        :param ambient_air_temp: Ambient air temperature (C)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Thermal conductivity of air (W/m*C or W/ft*C)
        """
        kf = None
        t_film = (conductor_temp + ambient_air_temp) / 2

        if uc.units_lookup[calculation_units] == uc.metric_value:
            kf = 2.424 * 10 ** -2 + 7.477 * 10 ** -5 * t_film - 4.407 * 10 ** -9 * t_film ** 2
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            kf = 0.007388 + 2.279 * 10 ** -5 * t_film - 1.343 * 10 ** -9 * t_film ** 2
        if df is not None:
            df.at[_idx, 'kf'] = kf
        return kf

    @staticmethod
    def c_pf(calculation_units, conductor_temp, ambient_air_temp, elevation, df=None, _idx=0):
        """
        Calculates air density and returns results (kg/m^3 or lb/ft^3)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param conductor_temp: Conductor temperature (C)
        :param ambient_air_temp: Ambient air temperature (C)
        :param elevation: elevation of conductors
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Air density (kg/m^3 or lb/ft^3
        """
        pf = None
        t_film = (conductor_temp + ambient_air_temp) / 2
        if uc.units_lookup[calculation_units] == uc.metric_value:
            pf = (1.293 - 1.525 * 10 ** -4 * elevation + 6.379 * 10 ** -9 * elevation ** 2) / (
                    1 + 0.00367 * t_film)
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            pf = (0.080695 - (2.901 * 10 ** -6) * elevation + (3.7 * 10 ** -11) * (elevation ** 2)) / (
                    1 + 0.00367 * t_film)
        if df is not None:
            df.at[_idx, 'pf'] = pf
        return pf

    @staticmethod
    def c_cond_resistance(conductor_temp, conductor_resistance, df=None, _idx=0):
        """
        Calculates resistance of conductor and returns results (ohm)
        :param conductor_temp: Conductor temperature (C)
        :param conductor_resistance: Dataframe containing resistance values/temperatures/distance
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Resistance (ohm)
        """
        high_resistance_ohm_per_unit_distance = conductor_resistance.at[0, 'high resistance Ω/unit']
        low_resistance_ohm_per_unit_distance = conductor_resistance.at[0, 'low resistance Ω/unit']
        high_resistance_temperature = conductor_resistance.at[0, 'high resistance temperature']
        low_resistance_temperature = conductor_resistance.at[0, 'low resistance temperature']
        resistance_distance = conductor_resistance.at[0, 'resistance distance']
        high_resistance = high_resistance_ohm_per_unit_distance / resistance_distance
        low_resistance = low_resistance_ohm_per_unit_distance / resistance_distance

        # todo add mention in report that conductor temperature is higher than resistance temperature and the results
        #  are less conservative
        #  overall resistance at temperature might be lower than what actually occurs physically
        #  reference page 10 738-2006
        resistance = (
                ((high_resistance - low_resistance) /
                 (high_resistance_temperature - low_resistance_temperature)) *
                (conductor_temp - low_resistance_temperature) + low_resistance
        )
        if df is not None:
            df.at[_idx, 'conductor resistance'] = resistance

        return resistance

    def c_Qs(self, calculation_units, atmosphere, latitude, day, month, year, hour, df=None, _idx=0):
        """
        Calculates total solar and sky radiated heat flux rate (W/m^2) and returns results
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Total solar and sky radiated heat flux rate (W/m^2)
        """
        aa = 0
        bb = 0
        cc = 0
        dd = 0
        ee = 0
        ff = 0
        gg = 0

        if uc.units_lookup[calculation_units] == uc.metric_value:
            if atmosphere == 'clear':
                # Clear atmosphere
                aa = -42.2391
                bb = 63.8044
                cc = -1.9220
                dd = 3.46921E-2
                ee = -3.61118E-4
                ff = 1.94318E-6
                gg = -4.07608E-9
            else:
                # Industrial atmosphere
                aa = 53.1821
                bb = 14.2110
                cc = 6.6138E-1
                dd = -3.1658E-2
                ee = 5.4654E-4
                ff = -4.3446E-6
                gg = 1.3236E-8
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            if atmosphere == 'clear':
                # Clear atmosphere
                aa = -3.9241
                bb = 5.9276
                cc = -1.7856E-1
                dd = 3.223E-3
                ee = -3.3549E-5
                ff = 1.8053E-7
                gg = -3.7868E-10
            else:
                # Industrial atmosphere
                aa = 4.9408
                bb = 1.3202
                cc = 6.1444E-2
                dd = -2.9411E-3
                ee = 5.07752E-5
                ff = -4.03627E-7
                gg = 1.22967E-9

        solar_altitude = self.c_solar_altitude(latitude, day, month, year, hour, df, _idx)
        radiated_heat_flux_rate = aa + bb * solar_altitude + cc * solar_altitude ** 2 + dd * solar_altitude ** 3 + ee * solar_altitude ** 4 + ff * solar_altitude ** 5 + gg * solar_altitude ** 6

        if df is not None:
            df.at[_idx, 'Qs'] = radiated_heat_flux_rate

        return radiated_heat_flux_rate

    def c_qse(self, calculation_units, elevation, atmosphere, latitude, day, month, year, hour, df=None, _idx=0):
        """
        Calculates elevation corrected total solar and sky radiated heat flux rate (W/m^2) and returns results
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param elevation: elevation of conductors
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: elevation corrected total solar and sky radiated heat flux rate (W/m^2)
        """
        k_solar = self.c_ksolar(calculation_units, elevation, df, _idx)
        Qs = self.c_Qs(calculation_units, atmosphere, latitude, day, month, year, hour, df, _idx)
        Qse = k_solar * Qs

        if df is not None:
            df.at[_idx, 'Qse'] = Qse

        return Qse

    @staticmethod
    def c_day_of_year(day, month, year, df, _idx):
        """
        Calculates day of year and returns results (int), ex: Jan 21 day_of_year = 21, Feb 12 day_of_year = 43
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Day of year (int)
        """
        error = False
        day_of_year = 0
        try:
            date = datetime.datetime(int(year), int(month), int(day))
            day_of_year = int(date.strftime("%j"))  # Get the day of the year
        except ValueError:  # if date is not valid
            df.at[_idx, 'Error'] = df.at[_idx, 'Error'].astype(
                str) + ' c_day_of_year: Check day/month/year, using 06/10/2009 '
            date = datetime.datetime(2009, 6, 10)
            day_of_year = int(date.strftime("%j"))  # Get the day of the year
            df.at[_idx, 'day of year'] = day_of_year
            error = True

        if df is not None and not error:
            df.at[_idx, 'day of year'] = day_of_year
        return day_of_year

    @staticmethod
    def c_ksolar(calculation_units, elevation, df=None, _idx=0):
        """
        Calculates solar heat multiplying factor (kSolar) and returns results
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param elevation: elevation of conductors
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Solar Heat multiplying factor
        """
        aks = 1.0
        bks = 3.500E-5
        cks = -1.000E-9
        solar_heat_factor = None

        if uc.units_lookup[calculation_units] == uc.metric_value:
            # meters
            if elevation < 1000:
                solar_heat_factor = 1.0
            elif 1000 <= elevation < 2000:
                solar_heat_factor = 1.10
            elif 2000 <= elevation < 4000:
                solar_heat_factor = 1.19
            elif 4000 <= elevation:
                solar_heat_factor = 1.28
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            # feet
            if elevation < 5000:
                solar_heat_factor = 1.0
            elif 5000 <= elevation < 10000:
                solar_heat_factor = 1.15
            elif 10000 <= elevation < 15000:
                solar_heat_factor = 1.25
            elif 15000 <= elevation:
                solar_heat_factor = 1.30

        solar_altitude_correction_factor = aks + bks * elevation + cks * elevation ** 2

        if df is not None:
            df.at[_idx, 'solar altitude correction factor'] = solar_heat_factor

        return solar_altitude_correction_factor

    @staticmethod
    def c_chi(omega, latitude, delta, df=None, _idx=0):
        """
        Calculates solar azimuth variable returns value (no units), used to calculate solar azimuth
        :param omega: hour angle (radians)
        :param latitude: latitude
        :param delta: solar declination (radians)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Solar Azimuth variable (no units)
        """
        chi = np.sin(omega) / (np.sin(latitude) * np.cos(omega) - np.cos(latitude) * np.tan(delta))
        if df is not None:
            df.at[_idx, 'chi'] = chi
        return chi

    @staticmethod
    def c_delta(day_of_year, df=None, _idx=0):
        """
        Calculates solar declination and returns results (radians)
        :param day_of_year: Day of Year, ex: Jan 21 day_of_year = 21, Feb 12 day_of_year = 43
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Solar declination angle (radians)
        """
        delta = np.radians(23.4583 * np.sin(np.radians((284 + day_of_year) / 365 * 360)))
        if df is not None:
            df.at[_idx, 'delta'] = delta
        return delta

    @staticmethod
    def c_omega(hour, df=None, _idx=0):
        """
        Calculates hour angle and returns results (radians)
        :param hour: Hour of day (int)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: hour angle (radians)
        """
        omega = np.radians((hour / 100 - 12) * 15)
        if df is not None:
            df.at[_idx, 'omega'] = omega
        return omega

    @staticmethod
    def c_solar_constant(omega, chi, df=None, _idx=0):
        """
        Calculates solar azimuth constant (C) and returns results (degrees)
        :param omega: hour angle (radians)
        :param chi: Solar Azimuth variable (no units)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Solar azimuth constant (degrees)
        """
        omega = np.degrees(omega)
        if -180 <= omega < 0:
            if chi >= 0:
                solar_azimuth_constant = 0
            else:
                solar_azimuth_constant = 180
        else:
            if chi >= 0:
                solar_azimuth_constant = 180
            else:
                solar_azimuth_constant = 360
        if df is not None:
            df.at[_idx, 'solar azimuth constant'] = solar_azimuth_constant
        return solar_azimuth_constant

    def c_solar_azimuth(self, latitude, day, month, year, hour, df=None, _idx=0):
        """
        Calculates solar azimuth and returns results (degrees)
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Solar Azimuth (degrees)
        """
        latitude = np.radians(latitude)
        day_of_year = self.c_day_of_year(day, month, year, df, _idx)
        delta = self.c_delta(day_of_year, df, _idx)
        omega = self.c_omega(hour, df, _idx)
        chi = self.c_chi(omega, latitude, delta, df, _idx)
        solar_constant = self.c_solar_constant(omega, chi, df, _idx)
        solar_azimuth = solar_constant + np.degrees(np.arctan(chi))
        if df is not None:
            df.at[_idx, 'solar azimuth'] = solar_azimuth
        return solar_azimuth

    def c_solar_altitude(self, latitude, day, month, year, hour, df=None, _idx=0):
        """
        Calculates solar altitude (Hc) and returns results (degrees)
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Solar altitude (degrees)
        """
        latitude = np.radians(latitude)
        day_of_year = self.c_day_of_year(day, month, year, df, _idx)
        delta = self.c_delta(day_of_year, df, _idx)
        omega = self.c_omega(hour, df, _idx)
        solar_altitude = np.degrees(np.arcsin(np.cos(latitude) * np.cos(delta) * np.cos(omega) +
                                              np.sin(latitude) * np.sin(delta)))
        if df is not None:
            df.at[_idx, 'hc: solar altitude'] = solar_altitude
        return solar_altitude

    def c_Theta(self, latitude, day, month, year, hour, conductor_direction, df=None, _idx=0):
        """
        Calculates Effective angles of incidence of the Sun's rays (θ) and returns results(degrees)
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param conductor_direction: Direction conductors run 'North/South' or 'East/West' (string)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Effective angles of incidence of the Sun's rays (degrees)
        """
        direction_lookup_value_ns = 1
        direction_lookup_value_ew = 2

        direction_lookup = {
            'n/s': direction_lookup_value_ns,
            'n-s': direction_lookup_value_ns,
            'north/south': direction_lookup_value_ns,
            'north-south': direction_lookup_value_ns,
            's/n': direction_lookup_value_ns,
            's-n': direction_lookup_value_ns,
            'south/north': direction_lookup_value_ns,
            'South-north': direction_lookup_value_ns,

            'e/w': direction_lookup_value_ew,
            'e-w': direction_lookup_value_ew,
            'east/west': direction_lookup_value_ew,
            'east-west': direction_lookup_value_ew,
            'w/e': direction_lookup_value_ew,
            'w-e': direction_lookup_value_ew,
            'west/east': direction_lookup_value_ew,
            'west-east': direction_lookup_value_ew,
        }

        if direction_lookup[conductor_direction.lower()] == direction_lookup_value_ns:
            _Z1 = 90
        else:
            _Z1 = 0

        solar_altitude = self.c_solar_altitude(latitude, day, month, year, hour, df, _idx)
        solar_azimuth = self.c_solar_azimuth(latitude, day, month, year, hour, df, _idx)

        theta = np.degrees(np.arccos(np.cos(np.radians(solar_altitude)) * np.cos(np.radians(solar_azimuth - _Z1))))

        if df is not None:
            df.at[_idx, 'theta'] = theta

        return theta

    @staticmethod
    def c_k_angle(wind_angle, df=None, _idx=0):
        """
        Calculates wind direction factor and returns results (no units)
        :param wind_angle: Angle between conductor and applied wind (degrees)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Wind direction factor (no units)
        """
        k_angle = 1.194 - np.cos(np.radians(wind_angle)) + 0.194 * np.cos(np.radians(2 * wind_angle)) + 0.368 * \
                  np.sin(np.radians(2 * wind_angle))
        if df is not None:
            df.at[_idx, 'k angle'] = k_angle
        return k_angle

    def c_qsHeatGain(self, calculation_units, solar_absorptivity, elevation, atmosphere, latitude, day, month,
                     year, hour, conductor_direction, conductor_projection, df=None, _idx=0):
        """
        Calculates heat gain rate from the sun and returns results (W/m or W/ft)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param solar_absorptivity: Solar absorptivity
        :param elevation: elevation of conductors
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param conductor_direction: Direction conductors run 'North/South' or 'East/West' (string)
        :param conductor_projection: Projected area of the conductor per unit length (diameter / 1000 or diameter / 12)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return:
        """
        qse = self.c_qse(calculation_units, elevation, atmosphere, latitude, day, month, year, hour, df, _idx)
        theta = self.c_Theta(latitude, day, month, year, hour, conductor_direction, df, _idx)
        qs_heat_gain = solar_absorptivity * qse * np.sin(np.radians(theta)) * conductor_projection

        if df is not None:
            df.at[_idx, 'qs heat gain'] = qs_heat_gain
        return qs_heat_gain

    @staticmethod
    def c_qrHeatLoss(calculation_units, diameter, emissivity, conductor_temp, ambient_air_temp, df=None, _idx=0):
        """
        Calculates radiated heat loss rate per unit length and returns results (W/m or W/ft)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param diameter: Conductor diameter (mm or in)
        :param emissivity: Emissivity
        :param conductor_temp: Conductor temperature (C)
        :param ambient_air_temp: Ambient air temperature (C)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Radiated heat loss rate per unit length (W/m or W/ft)
        """
        qr = None
        if uc.units_lookup[calculation_units] == uc.metric_value:
            # W/meters
            qr = 0.0178 * diameter * emissivity * (
                    ((conductor_temp + 273.15) / 100) ** 4 - ((ambient_air_temp + 273.15) / 100) ** 4)
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            # W/feet
            qr = 0.138 * diameter * emissivity * (
                    ((conductor_temp + 273.15) / 100) ** 4 - ((ambient_air_temp + 273.15) / 100) ** 4)

        if df is not None:
            df.at[_idx, 'qr heat loss'] = qr
        return qr

    def c_qcHeatLoss(self, calculation_units, diameter, conductor_temp, ambient_air_temp, elevation, wind_angle,
                     wind_speed, df=None, _idx=0):
        """
        Calculates convected heat loss rate per unit length and returns results (W/m or W/ft)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param diameter: Conductor diameter (mm or in)
        :param conductor_temp: Conductor temperature (C)
        :param ambient_air_temp: Ambient air temperature (C)
        :param elevation: elevation of conductors
        :param wind_angle: Angle between conductor and applied wind (degrees)
        :param wind_speed: Wind speed (m/s or ft/hr)
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Convected heat loss (W/m or W/ft)
        """
        qc0 = None
        qc1 = None
        qc2 = None
        qc_heat_loss = None
        k_angle = self.c_k_angle(wind_angle, df, _idx)
        uf = self.c_uf(calculation_units, conductor_temp, ambient_air_temp, df, _idx)
        kf = self.c_kf(calculation_units, conductor_temp, ambient_air_temp, df, _idx)
        pf = self.c_pf(calculation_units, conductor_temp, ambient_air_temp, elevation, df, _idx)

        if uc.units_lookup[calculation_units] == uc.metric_value:
            # W/meter
            # natural convection
            qc0 = 0.0205 * pf ** 0.5 * diameter ** 0.75 * (conductor_temp - ambient_air_temp) ** 1.25

            # low wind speeds
            qc1 = (1.01 + 0.0372 * ((diameter * pf * wind_speed) / uf) ** 0.52) * kf * k_angle * (
                    conductor_temp - ambient_air_temp)

            # high wind speeds
            qc2 = (0.0119 * ((diameter * pf * wind_speed) / uf) ** 0.6) * kf * k_angle * (
                    conductor_temp - ambient_air_temp)
            qc_heat_loss = np.amax((qc0, qc1, qc2))
        elif uc.units_lookup[calculation_units] == uc.imperial_value:
            # W/feet

            # natural convection
            qc0 = 0.283 * pf ** 0.5 * diameter ** 0.75 * (conductor_temp - ambient_air_temp) ** 1.25

            # low wind speeds
            qc1 = (1.01 + 0.371 * ((diameter * pf * wind_speed) / uf) ** 0.52) * kf * k_angle * (
                    conductor_temp - ambient_air_temp)

            # high wind speeds
            qc2 = (0.1695 * ((diameter * pf * wind_speed) / uf) ** 0.6) * kf * k_angle * (
                    conductor_temp - ambient_air_temp)
            qc_heat_loss = np.amax((qc0, qc1, qc2))

        if df is not None:
            df.at[_idx, 'qc0'] = qc0
            df.at[_idx, 'qc1'] = qc1
            df.at[_idx, 'qc2'] = qc2
            df.at[_idx, 'qc heat loss'] = qc_heat_loss
        return qc_heat_loss

    def c_SSRating(self, calculation_units, diameter, conductor_temp, ambient_air_temp, elevation, wind_angle,
                   wind_speed, emissivity, solar_absorptivity, atmosphere, latitude, day, month, year, hour,
                   conductor_direction, conductor_projection, conductor_resistance, df=None, _idx=0):
        """
        Calculates steady state current and returns results (Amps)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param diameter: Conductor diameter (mm or in)
        :param conductor_temp: Conductor temperature (C)
        :param ambient_air_temp: Ambient air temperature (C)
        :param elevation: elevation of conductors
        :param wind_angle: Angle between conductor and applied wind (degrees)
        :param wind_speed: Wind speed (m/s or ft/hr)
        :param emissivity: Emissivity
        :param solar_absorptivity: Solar absorptivity
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param conductor_direction: Direction conductors run 'North/South' or 'East/West' (string)
        :param conductor_projection: Projected area of the conductor per unit length (diameter / 1000 or diameter / 12)
        :param conductor_resistance: Dataframe containing resistance values/temperatures/distance
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return: Steady state current (Amps) Day rating includes solar heat gain, Night Rating does not include
        solar heat gain
        """
        qc = self.c_qcHeatLoss(calculation_units, diameter, conductor_temp, ambient_air_temp, elevation, wind_angle,
                               wind_speed, df, _idx)
        qr = self.c_qrHeatLoss(calculation_units, diameter, emissivity, conductor_temp, ambient_air_temp, df, _idx)
        qs = self.c_qsHeatGain(calculation_units, solar_absorptivity, elevation, atmosphere, latitude, day, month, year,
                               hour, conductor_direction, conductor_projection, df, _idx)
        r_cond = self.c_cond_resistance(conductor_temp, conductor_resistance, df, _idx)
        rating_day = self.current_steady_state(qr, qs, qc, r_cond)
        rating_night = self.current_steady_state(qr, 0, qc, r_cond) # solar heat gain not included for nighttime rating

        if df is not None:
            df.at[_idx, 'rating daytime'] = rating_day
            df.at[_idx, 'rating nighttime'] = rating_night
        return rating_day, rating_night

    @staticmethod
    def c_mcp(df, _idx):
        """
         Calculates conductor heat capacity based on material composition of the conductor
        :param df: Dataframe holding output conductor/config/calculated values
        :param _idx: index (row) for dataframe
        :return:
        """

        al = df.at[_idx, 'Al Weight'] / 1000
        cu = df.at[_idx, 'Cu Weight'] / 1000
        stl = df.at[_idx, 'St Weight'] / 1000
        alw = df.at[_idx, 'Alw Weight'] / 1000

        mcp_units = df.at[_idx, 'specific heat weight unit']

        if mcp_units == 'J/(kg-C)':
            # J/(kg-C)
            results = 955 * al + 423 * cu + 476 * stl + 534 * alw
        elif mcp_units == 'J/(lb-C)':
            # J/(lb-C)
            results = 433 * al + 192 * cu + 216 * stl + 242 * alw
        else:
            results = 0

        return results

    def c_initial_temp(self, calculation_units, diameter, conductor_temp_normal, conductor_temp_emergency,
                       ambient_air_temp, elevation, wind_angle, wind_speed, emissivity, solar_absorptivity,
                       atmosphere, latitude, day, month, year, hour, conductor_direction, conductor_projection,
                       conductor_resistance):
        """
        Calculates initial conductor temperatures for day and night ratings

        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param diameter: Conductor diameter (mm or in)
        :param conductor_temp_normal: conductor normal ampacity rating (amps)
        :param conductor_temp_emergency: conductor emergency ampacity rating (amps)
        :param ambient_air_temp: Ambient air temperature (C)
        :param elevation: elevation of conductors
        :param wind_angle: Angle between conductor and applied wind (degrees)
        :param wind_speed: Wind speed (m/s or ft/hr)
        :param emissivity: Emissivity
        :param solar_absorptivity: Solar absorptivity
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param conductor_direction: Direction conductors run 'North/South' or 'East/West' (string)
        :param conductor_projection: Projected area of the conductor per unit length (diameter / 1000 or diameter / 12)
        :param conductor_resistance: Dataframe containing resistance values/temperatures/distance
        :return:
        """

        initial_current_day, initial_current_night = self.c_SSRating(calculation_units, diameter, conductor_temp_normal,
                                                                     ambient_air_temp, elevation, wind_angle, 0,
                                                                     emissivity, solar_absorptivity, atmosphere,
                                                                     latitude, day, month, year, hour,
                                                                     conductor_direction, conductor_projection,
                                                                     conductor_resistance)

        result_day = optimize.minimize_scalar(self.c_find_initial_temp,
                                              bounds=(ambient_air_temp, conductor_temp_emergency),
                                              method='bounded',
                                              args=(
                                                  calculation_units, diameter, ambient_air_temp, elevation, wind_angle,
                                                  wind_speed, emissivity, solar_absorptivity, atmosphere, latitude,
                                                  day, month, year, hour, conductor_direction, conductor_projection,
                                                  conductor_resistance, initial_current_day, 'Day'))

        result_night = optimize.minimize_scalar(self.c_find_initial_temp,
                                                bounds=(ambient_air_temp, conductor_temp_emergency),
                                                method='bounded',
                                                args=(
                                                    calculation_units, diameter, ambient_air_temp, elevation,
                                                    wind_angle, wind_speed, emissivity, solar_absorptivity, atmosphere,
                                                    latitude, day, month, year, hour, conductor_direction,
                                                    conductor_projection, conductor_resistance, initial_current_night,
                                                    "Night"))

        return result_day.x, result_night.x

    def c_find_initial_temp(self, conductor_temperature, calculation_units, diameter, ambient_air_temp, elevation, wind_angle,
                            conductor_wind_emergency, emissivity, solar_absorptivity, atmosphere, latitude, day, month,
                            year,
                            hour, conductor_direction, conductor_projection, conductor_resistance, initial_current,
                            condition):
        """
        Helper function that returns the difference between initial_current and current_rating_day/current_rating_night
        conductor_temperature varied by c_initial_temp with the goal finding a conductor_temperature that makes
        current_rating_day/current_rating_night = initial_current (condition = 'Day' or 'Night')

        :param conductor_wind_emergency: Emergency rating wind speed (m/s or ft/hr)
        :param conductor_temperature: conductor temperature (C)
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param diameter: Conductor diameter (mm or in)
        :param ambient_air_temp: Ambient air temperature (C)
        :param elevation: elevation of conductors
        :param wind_angle: Angle between conductor and applied wind (degrees)
        :param emissivity: Emissivity
        :param solar_absorptivity: Solar absorptivity
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param conductor_direction: Direction conductors run 'North/South' or 'East/West' (string)
        :param conductor_projection: Projected area of the conductor per unit length (diameter / 1000 or diameter / 12)
        :param conductor_resistance: Dataframe containing resistance values/temperatures/distance
        :param initial_current: initial current of conductor
        :param condition: Day/Night
        :return:
        """

        delta = 0 #
        current_rating_day, current_rating_night = self.c_SSRating(calculation_units, diameter, conductor_temperature, ambient_air_temp,
                                                                   elevation, wind_angle, conductor_wind_emergency,
                                                                   emissivity, solar_absorptivity, atmosphere, latitude,
                                                                   day, month, year, hour, conductor_direction,
                                                                   conductor_projection, conductor_resistance)
        if condition == 'Day':
            delta = initial_current - current_rating_day
        elif condition == 'Night':
            delta = initial_current - current_rating_night

        delta = np.abs(delta)
        return delta

    def find_conductor_temp(self, conductor_temperature, calculation_units, diameter, ambient_air_temp, elevation, wind_angle,
                            conductor_wind_emergency, emissivity, solar_absorptivity, atmosphere, latitude, day, month,
                            year, hour, conductor_direction, conductor_projection, conductor_resistance,
                            initial_temperature, initial_current, conductor_temp_emergency, mcp, condition, duration):
        """
        Calculates initial conductor temperature
        Takes normal conductor rating and applies
        :param conductor_temperature:
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param diameter: Conductor diameter (mm or in)
        :param ambient_air_temp: Ambient air temperature (C)
        :param elevation: elevation of conductors
        :param wind_angle: Angle between conductor and applied wind (degrees)
        :param conductor_wind_emergency: Emergency rating wind speed (m/s or ft/hr)
        :param emissivity: Emissivity
        :param solar_absorptivity: Solar absorptivity
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param conductor_direction: Direction conductors run 'North/South' or 'East/West' (string)
        :param conductor_projection: Projected area of the conductor per unit length (diameter / 1000 or diameter / 12)
        :param conductor_resistance: Dataframe containing resistance values/temperatures/distance
        :param initial_temperature: conductor initial temperature (C)
        :param initial_current: initial current of conductor
        :param conductor_temp_emergency: conductor emergency ampacity rating (amps)
        :param mcp:  conductor heat capacity (lb-C)
        :param condition: Day/Night
        :param duration: time frame for increased current rating (seconds)
        :return: returns conductor temperature (C)
        """

        final_ = 0

        # TODO can this be done better? why did I do this?
        if condition == 'Day':
            final_, _ = self.c_SSRating(calculation_units, diameter, conductor_temperature, ambient_air_temp,
                                        elevation, wind_angle, conductor_wind_emergency, emissivity, solar_absorptivity,
                                        atmosphere, latitude, day, month, year, hour, conductor_direction,
                                        conductor_projection, conductor_resistance)
        elif condition == 'Night':
            _, final_ = self.c_SSRating(calculation_units, diameter, conductor_temperature, ambient_air_temp,
                                        elevation, wind_angle, conductor_wind_emergency, emissivity, solar_absorptivity,
                                        atmosphere, latitude, day, month, year, hour, conductor_direction,
                                        conductor_projection, conductor_resistance)

        if self.true_to_standard:
            # todo make sure this is correct
            # returns conductor resistance at 162% of conductor initial temperature
            r = self.c_cond_resistance(initial_temperature * 1.62, conductor_resistance)
        else:
            # returns conductor resistance at conductor initial temperature
            r = self.c_cond_resistance(initial_temperature, conductor_resistance)

        calc_tau = (mcp * (conductor_temperature - initial_temperature)) / (r * (final_ ** 2 - initial_current ** 2)) * 1 / 60

        tc = (initial_temperature + (conductor_temperature - initial_temperature) * (1 - np.exp(-duration / calc_tau)))
        delta = np.abs(conductor_temp_emergency - tc)
        return delta

    def load_dump(self, calculation_units, diameter, conductor_temp_normal, conductor_temp_emergency,
                  ambient_air_temp, elevation, wind_angle, wind_speed, emissivity, solar_absorptivity,
                  atmosphere, latitude, day, month, year, hour, conductor_direction, conductor_projection,
                  conductor_resistance, mcp, duration):
        """
        The secret sauce of all of this.
        Calculates the maximum current through a conductor for a given set of initial conditions over a specified
        duration that is limited by a maximum conductor temperature
        :param calculation_units: Units: 'Metric' or 'Imperial'
        :param diameter: Conductor diameter (mm or in)
        :param conductor_temp_normal: conductor normal ampacity rating (amps)
        :param conductor_temp_emergency: conductor emergency ampacity rating (amps)
        :param ambient_air_temp: Ambient air temperature (C)
        :param elevation: elevation of conductors
        :param wind_angle: Angle between conductor and applied wind (degrees)
        :param wind_speed: Wind speed (m/s or ft/hr)
        :param emissivity: Emissivity
        :param solar_absorptivity: Solar absorptivity
        :param atmosphere: Atmospheric conditions 'Industrial' or 'Clear'
        :param latitude: latitude
        :param day: Day of month (int)
        :param month: Month (int)
        :param year: Year (int)
        :param hour: Hour of day (int)
        :param conductor_direction: Direction conductors run 'North/South' or 'East/West' (string)
        :param conductor_projection: Projected area of the conductor per unit length (diameter / 1000 or diameter / 12)
        :param conductor_resistance: Dataframe containing resistance values/temperatures/distance
        :param mcp: conductor heat capacity (lb-C)
        :param duration: time frame for increased current rating (seconds)
        :return: current (A)
        """
        # Calculate steady state current for daytime & nighttime rating with 0 wind
        initial_current_day, initial_current_night = \
            self.c_SSRating(calculation_units, diameter, conductor_temp_normal,
                            ambient_air_temp, elevation, wind_angle, 0, emissivity, solar_absorptivity, atmosphere,
                            latitude, day, month, year, hour, conductor_direction, conductor_projection,
                            conductor_resistance)

        # Calculate initial conductor temperature for daytime & nighttime rating with emergency wind applied
        initial_temperature_day, initial_temperature_night = \
            self.c_initial_temp(calculation_units, diameter, conductor_temp_normal, conductor_temp_emergency,
                                ambient_air_temp, elevation, wind_angle, wind_speed, emissivity, solar_absorptivity,
                                atmosphere, latitude, day, month, year, hour, conductor_direction,
                                conductor_projection, conductor_resistance)

        # Use scipy.minimize find the maximum daytime conductor temperature
        # Adjusts find_conductor_temp variable between ambient_air_temp & 600 C)
        result_day = optimize.minimize_scalar(self.find_conductor_temp, bounds=(ambient_air_temp, 600),
                                              method='bounded',
                                              args=(
                                                  calculation_units, diameter, ambient_air_temp, elevation, wind_angle,
                                                  wind_speed, emissivity, solar_absorptivity, atmosphere, latitude,
                                                  day, month, year, hour, conductor_direction, conductor_projection,
                                                  conductor_resistance, initial_temperature_day, initial_current_day,
                                                  conductor_temp_emergency, mcp, 'Day', duration))

        # Use scipy.minimize find the maximum nighttime conductor temperature
        # Adjusts find_conductor_temp variable between ambient_air_temp & 600 C)
        result_night = optimize.minimize_scalar(self.find_conductor_temp, bounds=(ambient_air_temp, 600),
                                                method='bounded',
                                                args=(
                                                    calculation_units, diameter, ambient_air_temp, elevation,
                                                    wind_angle, wind_speed, emissivity, solar_absorptivity, atmosphere,
                                                    latitude, day, month, year, hour, conductor_direction,
                                                    conductor_projection, conductor_resistance,
                                                    initial_temperature_night, initial_current_night,
                                                    conductor_temp_emergency, mcp, 'Night', duration))

        # optimize.minimize_scalar returns more than just a value, .x returns desired values
        final_temperature_day = result_day.x
        final_temperature_night = result_night.x

        # Calculates maximum daytime rating based on maximum conductor temperature found by optimize.minimize_scalar
        final_current_day, _ = self.c_SSRating(calculation_units, diameter, final_temperature_day, ambient_air_temp,
                                               elevation, wind_angle, wind_speed, emissivity, solar_absorptivity,
                                               atmosphere, latitude, day, month, year, hour, conductor_direction,
                                               conductor_projection, conductor_resistance)

        # Calculates maximum nighttime rating based on maximum conductor temperature found by optimize.minimize_scalar
        # Nighttime rating requires additional call due to the fact that final_temperature_night ≠ final_temperature_day
        _, final_current_night = self.c_SSRating(calculation_units, diameter, final_temperature_night, ambient_air_temp,
                                                 elevation, wind_angle, wind_speed, emissivity, solar_absorptivity,
                                                 atmosphere, latitude, day, month, year, hour, conductor_direction,
                                                 conductor_projection, conductor_resistance)

        return final_current_day, final_current_night
